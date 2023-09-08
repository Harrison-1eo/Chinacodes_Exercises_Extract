"""
 @project: storage.py
 @description: 定义存储答案函数，将答案存储到Excel表格中，方便后续查询和修改
 @author: Harrison-1eo
 @date: 2023-09-08
 @version: 1.0
"""

import openpyxl
import os

# 查找是否存在Excel表，不存在则创建
def create_excel():
    if not os.path.exists('answer.xlsx'):
        wb = openpyxl.Workbook()
        wb.active.column_dimensions['A'].width = 100
        wb.active.column_dimensions['B'].width = 50
        wb.save('answer.xlsx')
        print("创建Excel表成功")
        
# 读取Excel表中的题目和答案，第一列是题目，第二列是答案
def read_excel():
    wb = openpyxl.load_workbook('answer.xlsx')
    ws = wb.active
    res = {}
    for row in ws.rows:
        res.update({row[0].value: row[1].value})
    return res

# 将题目和答案写入Excel表中
def write_excel(ans_dict: dict, answers: list):
    wb = openpyxl.load_workbook('answer.xlsx')
    ws = wb.active
    
    for question, answer in answers:
        if question not in ans_dict:
            ws.append([question, answer])
            ans_dict.update({question: answer})
    wb.save('answer.xlsx')
    print("写入Excel表成功")
    
    

    

    
